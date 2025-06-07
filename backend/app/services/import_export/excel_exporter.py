"""Excel export service for ARS data."""

import pandas as pd
from typing import Any, Dict, List
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .base_exporter import BaseExporter, ExportResult


class ExcelExporter(BaseExporter):
    """Excel format exporter for ARS data."""
    
    def export_reporting_event(self, reporting_event_id: str, output_path: str) -> ExportResult:
        """Export a complete reporting event to Excel format."""
        try:
            self._update_progress(0, 100, "Starting Excel export...")
            
            # Get the data
            data = self._get_reporting_event_data(reporting_event_id)
            self._update_progress(30, 100, "Data retrieved, validating...")
            
            # Validate data
            errors = self._validate_export_data(data)
            if errors:
                return ExportResult(
                    success=False,
                    message=f"Validation failed: {'; '.join(errors)}",
                    records_exported=0,
                    errors=errors,
                    export_time=datetime.now()
                )
            
            self._update_progress(50, 100, "Creating Excel workbook...")
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each data type
            sheet_mapping = {
                "ReportingEvent": "reportingEvent",
                "Analyses": "analyses",
                "Methods": "methods",
                "AnalysisSets": "analysisSets",
                "DataSubsets": "dataSubsets",
                "Groups": "groups",
                "Operations": "operations",
                "Outputs": "outputs",
                "WhereClauses": "whereClauses"
            }
            
            for sheet_name, data_key in sheet_mapping.items():
                if data.get(data_key):
                    self._create_excel_sheet(wb, sheet_name, data[data_key])
            
            # Add metadata sheet
            self._create_metadata_sheet(wb, reporting_event_id)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save workbook
            wb.save(output_path)
            
            self._update_progress(100, 100, "Export completed")
            
            return ExportResult(
                success=True,
                message="Excel export completed successfully",
                file_path=output_path,
                records_exported=len(data.get("analyses", [])),
                export_time=datetime.now()
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Excel export failed: {str(e)}",
                records_exported=0,
                errors=[str(e)],
                export_time=datetime.now()
            )
    
    def export_analyses(self, analysis_ids: List[str], output_path: str) -> ExportResult:
        """Export specific analyses to Excel format."""
        try:
            self._update_progress(0, 100, "Starting analyses export...")
            
            data = {"analyses": []}
            
            from app.models.ars import Analysis
            
            for i, analysis_id in enumerate(analysis_ids):
                analysis = self.db.query(Analysis).filter(Analysis.id == analysis_id).first()
                if analysis:
                    data["analyses"].append(self._serialize_model(analysis))
                    
                progress = ((i + 1) / len(analysis_ids)) * 80
                self._update_progress(progress, 100, f"Processing analysis {i + 1}/{len(analysis_ids)}")
            
            # Create Excel workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            self._create_excel_sheet(wb, "Analyses", data["analyses"])
            self._create_metadata_sheet(wb, None, analysis_ids)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save workbook
            wb.save(output_path)
            
            self._update_progress(100, 100, "Export completed")
            
            return ExportResult(
                success=True,
                message="Analyses Excel export completed successfully",
                file_path=output_path,
                records_exported=len(data["analyses"]),
                export_time=datetime.now()
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Analyses Excel export failed: {str(e)}",
                records_exported=0,
                errors=[str(e)],
                export_time=datetime.now()
            )
    
    def export_custom_query(self, query_params: Dict[str, Any], output_path: str) -> ExportResult:
        """Export data based on custom query parameters to Excel format."""
        try:
            self._update_progress(0, 100, "Processing custom query...")
            
            # Create Excel workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Create query results sheet
            ws = wb.create_sheet("QueryResults")
            ws.append(["Custom Query Results"])
            ws.append(["Query executed at:", datetime.now().isoformat()])
            ws.append([])
            
            # Add query parameters
            ws.append(["Query Parameters:"])
            for key, value in query_params.items():
                ws.append([key, str(value)])
            
            # Style the header
            header_font = Font(bold=True, size=14)
            ws['A1'].font = header_font
            
            # Create metadata sheet
            self._create_metadata_sheet(wb, None, None, query_params)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save workbook
            wb.save(output_path)
            
            self._update_progress(100, 100, "Custom query export completed")
            
            return ExportResult(
                success=True,
                message="Custom query Excel export completed successfully",
                file_path=output_path,
                records_exported=0,
                export_time=datetime.now()
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                message=f"Custom query Excel export failed: {str(e)}",
                records_exported=0,
                errors=[str(e)],
                export_time=datetime.now()
            )
    
    def _create_excel_sheet(self, workbook: Workbook, sheet_name: str, data: List[Dict[str, Any]]):
        """Create an Excel sheet with data."""
        if not data:
            return
            
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Create worksheet
        ws = workbook.create_sheet(sheet_name)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metadata_sheet(self, workbook: Workbook, reporting_event_id: str = None, 
                              analysis_ids: List[str] = None, query_params: Dict[str, Any] = None):
        """Create a metadata sheet with export information."""
        ws = workbook.create_sheet("Metadata")
        
        # Add metadata
        metadata = [
            ["Export Information"],
            [""],
            ["Export Date", datetime.now().isoformat()],
            ["Export Format", "Excel"],
            ["Exporter", "ARS Excel Exporter v1.0"],
            [""]
        ]
        
        if reporting_event_id:
            metadata.extend([
                ["Reporting Event ID", reporting_event_id],
                [""]
            ])
        
        if analysis_ids:
            metadata.extend([
                ["Analysis IDs", ", ".join(analysis_ids)],
                [""]
            ])
        
        if query_params:
            metadata.extend([
                ["Query Parameters"],
                [""]
            ])
            for key, value in query_params.items():
                metadata.append([key, str(value)])
        
        for row in metadata:
            ws.append(row)
        
        # Style the header
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font